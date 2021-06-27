# This files contains your custom actions which can be used to run
# custom Python code.
#
# See this guide on how to implement these action:
# https://rasa.com/docs/rasa/custom-actions


# This is a simple example for a custom action which utters "Hello World!"
from typing import Any, Text, Dict, List
from openpyxl.reader.excel import ExcelReader
from pandas.io.pytables import AppendableFrameTable
from rasa_sdk import Action, Tracker
from rasa_sdk.executor import CollectingDispatcher
from rasa_sdk.events import SlotSet,AllSlotsReset
#from policies.policy import TestPolicy 
import pandas as pd
from pandas import ExcelWriter
import openpyxl as op
from openpyxl import load_workbook



#
#
# class ActionHelloWorld(Action):
#
#     def name(self) -> Text:
#         return "action_hello_world"
#
#     def run(self, dispatcher: CollectingDispatcher,
#             tracker: Tracker,
#             domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
#
#         dispatcher.utter_message(text="Hello World!")
#
#         return []

class imprimirSlot(Action):
    def name(self) -> Text:
        return "action_imprimir_determinantes"
        
    def run(self, dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        ######CONTENIDOS######:
        
        #1 figura humana completa
        _H = {"persona","humano","hombre","mujer","niño","niña","chico","chica","señor","señora","personas","humanos","hombres","mujeres","niños","niñas","chicos","chicas","señores","señoras"}
        #2 figura humana completa irreal, de ficción o mitológica
        _ParentesisH = {"payaso","payasos","hada","hadas","bruja","brujas","fantasma","fantasmas","enano","enanos","enana","enanas","demonio","demonios","ángel","ángeles","humanoide","humanoides","caricaturas","caricatura","monstruo","monstruos","duende","duendes"}
        #3 Detalle humano
        _Hd = {"brazo","pierna","dedos","pies","cabeza","codo","nariz"}        
        #5 Experiencia humana
        _Hx = {"amor","amar","ama","amamos","amo","odio","odia","odiamos","odiar","depresión","deprimido","deprimida","deprimidos","deprimidas","feliz","felices","alegre","alegres","felicidad","alegria","ruido","ruidoso","sonido","suena","olor","huele","oloroso","miedo","temor","miedoso","contento","contenta","contentos","contentas"}
        #6 Figura animal completa
        _A = {"escarabajo","escarabajos","bicho","bichos","araña","arañas","cucaracha","cucarachas","mariposa","mariposas","mantis","mosca","mosquito","moscas","mosquitos","murcielago","murciélago","murciélagos","pulga","pulgas","águila","águilas","avestruz","ballena","bisonte","búfalo","búhos","buitre","burro","caballo","cabra","camaleón","camello","canario","castor","cebra","cerdo","chancho","ciervo","cobra","colibrí","comadreja","cóndor","conejo","delfín","elefante","faisan","flamenco","foca","gallina","gallo","gato","gorila","guepardo","hámster","hiena","hipopótamo","jabalí","jaguar","jirafa","koala","lagarto","león","lobo","loro","manatí","mapache","mono","murciélago","nutria","ñandú","orcas","oso","pájaro","paloma","panda","pato","pavo","pelícano","perro","pingüino","puercoespín","puma","rana","ratón","reno","rinoceronte","salamandra","sapo","serpiente","tapir","tejon","tiburón","tigre","topo","toro","tucán","vaca","vicuña","zorrino","zorro","águila","avestruces","ballenas","bisontes","búfalos","búho","buitres","burros","caballos","cabras","camaleones","camellos","canarios","castores","cebras","cerdos","chanchos","ciervos","cobras","colibries","comadrejas","cóndores","conejos","delfines","elefantes","faisanes","flamencos","focas","gallinas","gallos","gatos","gorilas","guepardos","hámsters","hienas","hipopótamos","jabalies","jaguares","jirafas","koalas","lagartos","leones","lobos","loros","manaties","mapaches","monos","murciélagos","nutrias","ñandues","orca","osos","pájaros","palomas","pandas","patos","pavos","pelícanos","perros","pingüinos","puercoespines","pumas","ranas","ratones","renos","rinocerontes","salamandras","sapos","serpientes","tapires","tiburones","tigres","topos","toros","tucanes","vacas","víbora","víboras","vicuñas","zorrinos","zorros"}
        #7 Figura animal completa irreal, de ficción o mitológica
        _ParentesisA = {"unicornio","unicornios","dragón","dragon","dragones","minotauro","minotauros"}
        #8 Figura animal incompleta
        _Ad = {"pata","patas","cola","pinza","hocico", "cuero","pezuñas","garras","vasos","pico"}
        #10 Anatomía
        _An = {"ósea","osea","cráneo","cráneo","torax","toracica","tórax","corazón","corazon","pulmón","pulmon","estomágo","estomago","panza","hígado","higado","musculo","articulaciones","vértebral","vértebra","vertebras","cerebro","cerebros"}
        #11 Arte
        _Art = {"pintura","dibujo","pinturas","dibujos","ilustración","ilustracion","ilustraciones","acuarela","acuarelas","arte","estatua","estatuas","escultura","esculturas","joya","joyas","insignia","insignias","escudo","escudos","adornos","espada","espadas","cuadros","lienzos","cuadro","lienzo"}
        #13 Sangre
        _Bl = {"sangre","sanguíneo","sanguínea","sanguinario","sanguinaria","sangriento","sangrienta"}
        #14 Botánica
        _Bt = {"vegetal","vegetales","arbusto","arbustos","flor","flores","floral","alga","algas","arbol","árbol","árboles","arboles","hoja","hojas","pétalo","pétalos","tronco","tronco","raiz","raíces","nido","hongo","hongos","pino","palmera","pinos","palmeras"} 
        #15 Vestidos
        _Cg = {"sombrero","sombreros","gorro","gorros","gorras","gorra","bota","botas","botines","cinturón","cinturon","cinturones","corbata","corbatas","moño","moños","chaqueta","chaquetas","saco","sacos","polera","poleras","pantalón","pantalon","pantalones","bufanda","bufandas","bermudas","short","shorts","medias","calzoncillos","musculosa"}
        #16 Nubes
        _Cl = {"nubes","nube","nubarrón","nubarrones","nublado","nublada"}
        #17 Explosión
        _Ex = {"explosión","explosión","explotar","implotar","bomba","explosivos","estallido","estallidos","detonación","detonaciones","bombardeo","bombardeos"}
        #18 Fuego
        _Fi = {"fuego","fuegos","llama","llamas","incendio","incendios","humo","humos","fogata","fogatas","chispa","chispas","hoguera","hogueras","quema","quemas"}
        #19 Comida
        _Fd = {"pollo","pollos","carne","carnes","pescado","pescados","helado","helados","panqueque","panqueques","verdura","verduras","papa","papas","zapallo","zapallos","lechuga","lechugas","tomates","tomate","zanahoria","zanahorias","sandía","sandías","melón","melones","kiwi","kiwis","mandarina","naranja","mandarinas","naranjas","banana","bananas","palta","frutas","verdura","fruta","torta","budín","tostadas","sandwich","sandwiches","pan","panes","sopa","pasta","sopas","pastas","spaghetti"}
        #20 Geografía
        _Ge = {"mapa","mapas","plano","planos","cartográfico","cartografía","continente","continentes","país","paises","ciudad","ciudad","region","regiones"}
        #21 Hogar
        _Hh = {"cama","cucheta","sommier","sillón","silla","mesa","lámpara","cuchillo","olla","alfombra","cortina","mueble","horno","cocina","pieza","habitación","baño","cochera","garage","camas","cuchetas","sommiers","sillones","sillas","mesas","lámparas","cuchillos","ollas","alfombras","cortinas","muebles","hornos","cocinas","piezas","habitaciones","baños","cocheras"}
        #22 Paisaje
        _Ls = {"montaña","montañas","cordillera","cordilleras","colina","colinas","cerro","cerros","sierra","sierras","isla","islas","cueva","cuevas","roca","rocas","piedra","piedras","bosque","bosques","desierto","desierto","llanura","llanuras","pantano","pantanos","glaciar","glaciares"}
        #23 Naturaleza
        _Na = {"lago","laguna","lagos","lagunas","sol","luna","mar","mar","océano","océanos","agua","hielo","hielos","lluvia","lluvias","niebla","neblina","bruma","tormenta","brisa","viento","arco iris","tormenta","noche","día","dia","granizo","trueno","relámpago","relampago"}
        #24 Ciencia
        _Sc = {"avión","aviones","avion","edificio","edificios","puente","puentes","auto","autos","moto","motos","microscopio","microscopios","laboratorio","laboratorios","motor","motores","turbina","turbinas","telescopio","telescopios","arma","armas","armamento","cohete","cohetes","nave","naves","ovni","OVNI","ovnis","OVNIS","barco","barcos","antena","antenas","satélite","satélites","satelite","satelites"}
        #25 Sexo
        _Sx = {"pene","penes","verga","vergas","pito","pitos","vagina","vaginas","concha","conchas","nalgas","cachas","pechos","teta","tetas","testículos","huevos","bolas","menstruación","aborto","abortar","coito","coger","garchar","cogiendo","teniendo sexo","garchando"}
        #26 Radiografía
        _Xy = {"radiografía","radiografia","placa","placas","rayos x"}
        #27 Popular1
        _Po1 = {"escarabajo","escarabajos","bicho","bichos","araña","arañas","cucaracha","cucarachas","mariposa","mariposas","mantis","mosca","mosquito","moscas","mosquitos","insecto","insectos","gusano"}
        #28 Popular3
        _Po3 = {"persona","humano","hombre","mujer","niño","niña","chico","chica","señor","señora","personas","humanos","hombres","mujeres","niños","niñas","chicos","chicas","señores","señoras","payaso","payasos","hada","hadas","bruja","brujas","fantasma","fantasmas","enano","enanos","enana","enanas","demonio","demonios","ángel","ángeles","humanoide","humanoides","caricaturas","caricatura","monstruo","monstruos","duende","duendes"}
        #IMAGEN 2 NO TIENE RESPUESTAS POPULARES
        determinantes = '?'
        par = ''
        contenidos = ''
        popular = ''
        lamina = tracker.get_slot("contador")
        respuesta = tracker.latest_message['text']
        dispatcher.utter_message(text="CONTENIDOS:")
        while _H:#1
            subconjunto = _H.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'H,'
                dispatcher.utter_message(text="Hay figura humana completa(contenido)")        
                break   
        while _ParentesisH:#2
            subconjunto = _ParentesisH.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + '(H),'
                dispatcher.utter_message(text="Hay figura completa irreal, de ficción o mitológica(contenido)")        
                break
        while _Hd: #3
            subconjunto = _Hd.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Hd,'
                dispatcher.utter_message(text="Hay detalle humano(contenido)")        
                break
        while _Hx: #5
            subconjunto = _Hx.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Hx,'
                dispatcher.utter_message(text="Hay experiencia humana(contenido)")        
                break
        while _A: #6
            subconjunto = _A.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'A,'
                dispatcher.utter_message(text="Hay figura animal completa(contenido)")        
                break
        while _ParentesisA: #7
            subconjunto = _ParentesisA.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + '(A),'
                dispatcher.utter_message(text="Hay figura animal completa irreal, de ficción o mitológica(contenido)")        
                break
        while _Ad: #8
            subconjunto = _Ad.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Ad,'
                dispatcher.utter_message(text="Hay figura animal incompleta(contenido)")        
                break
        while _An: #10
            subconjunto = _An.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'An,'
                dispatcher.utter_message(text="Hay anatomía(contenido)")        
                break
        while _Art: #11
            subconjunto = _Art.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Art,'
                dispatcher.utter_message(text="Hay arte(contenido)")        
                break
        while _Bl: #13
            subconjunto = _Bl.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Bl,'
                dispatcher.utter_message(text="Hay sangre(contenido)")        
                break
        while _Bt: #14
            subconjunto = _Bt.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Bt,'
                dispatcher.utter_message(text="Hay botánica(contenido)")        
                break
        while _Cg: #15
            subconjunto = _Cg.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Cg,'
                dispatcher.utter_message(text="Hay vestidos(contenido)")        
                break
        while _Cl: #16
            subconjunto = _Cl.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Cl,'
                dispatcher.utter_message(text="Hay nubes(contenido)")        
                break
        while _Ex: #17
            subconjunto = _Ex.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Ex,'
                dispatcher.utter_message(text="Hay explosión(contenido)")        
                break
        while _Fi: #18
            subconjunto = _Fi.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Fi,'
                dispatcher.utter_message(text="Hay fuego(contenido)")        
                break
        while _Fd: #19
            subconjunto = _Fd.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Fd,'
                dispatcher.utter_message(text="Hay comida(contenido)")        
                break
        while _Ge: #20
            subconjunto = _Ge.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Ge,'
                dispatcher.utter_message(text="Hay geografía(contenido)")        
                break
        while _Hh: #21
            subconjunto = _Hh.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Hh,'
                dispatcher.utter_message(text="Hay hogar(contenido)")        
                break
        while _Na: #23
            subconjunto = _Na.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Na,'
                dispatcher.utter_message(text="Hay naturaleza(contenido)")        
                break
        while _Sc: #24
            subconjunto = _Sc.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Sc,'
                dispatcher.utter_message(text="Hay ciencia(contenido)")        
                break
        while _Sx: #25
            subconjunto = _Sx.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Sx,'
                dispatcher.utter_message(text="Hay sexo(contenido)")        
                break
        while _Xy: #26
            subconjunto = _Xy.pop()
            if subconjunto in respuesta:
                contenidos = contenidos + 'Xy,'
                dispatcher.utter_message(text="Hay radiografía(contenido)")        
                break
        while _Po1: #27 populares
            subconjunto = _Po1.pop()
            if subconjunto in respuesta:
                if lamina == '1':#and zona corresponde
                    popular = 'Po1' #hay que revisar si coincide con sector de la imagen
                    dispatcher.utter_message(text="Hay respuesta popular de imagen 1(contenido)")
                    break
                else: 
                    popular = '?'        
            else: popular ='?'    

        #IMAGEN 2 NO TIENE RESPUESTAS POPULARES

        while _Po3: #28 populares
            subconjunto = _Po3.pop()
            if subconjunto in respuesta:
                if lamina == '3': #and zona corresponde
                    popular = 'Po3' 
                    dispatcher.utter_message(text="Hay respuesta popular de imagen 3(contenido)")        
                    break
                else: popular ='?'
            else: popular ='?'
        #####comprobar populares al final según orden de las respuestas

        ##todo lo que no entra en un conjunto es contenido ideográfico "Id"
        sender = tracker.sender_id


        ######DETERMINANTES#####
        slot_paridad = tracker.get_slot("par")
        slot_vista = tracker.get_slot("vista")
        slot_ccromatico = tracker.get_slot("color_cromatico")
        slot_cacromatico = tracker.get_slot("color_acromatico")
        slot_forma = tracker.get_slot("forma")
        slot_movimiento = tracker.get_slot("movimiento")
        slot_textura = tracker.get_slot("textura")
        slot_reflejo = tracker.get_slot("reflejo")       
        
        dispatcher.utter_message(text="DETERMINANTES:")
        if slot_paridad == "true":
            par = '2' 
            dispatcher.utter_message(text="Hay par(determinante)")
        else:
            par ='?'
            dispatcher.utter_message(text="No hay par(determinante)")
        if slot_vista == "true":
            if determinantes == '?':
                determinantes = 'V,'
            else: 
                determinantes = determinantes + 'V,'
            dispatcher.utter_message(text="Hay vista(determinante)")
        else:
            dispatcher.utter_message(text="No hay vista(determinante)")
        if slot_ccromatico == "true":
            if determinantes == '?':
                determinantes = 'C,'
            else: 
                determinantes = determinantes + 'C,'
            dispatcher.utter_message(text="Hay color cromático(determinante)")
        else:
            dispatcher.utter_message(text="No hay color cromático(determinante)")
        if slot_cacromatico == "true":
            if determinantes == '?':
                determinantes = 'C\','
            else: 
                determinantes = determinantes + 'C\','
            dispatcher.utter_message(text="Hay color acromático(determinante)")
        else:
            dispatcher.utter_message(text="No hay color acromático(determinante)")
        if slot_forma == "humana":
            if determinantes == '?':
                determinantes = 'M,'
            else: 
                determinantes = determinantes + 'M,'
            dispatcher.utter_message(text="Hay forma humana(determinante)")
        else:
            if slot_forma == "animal":
                if determinantes == '?':
                    determinantes = 'FM,'
                else: 
                    determinantes = determinantes + 'FM,'
                dispatcher.utter_message(text="Hay forma animal(determinante)")
            else:
                if slot_forma == "inanimada":
                    if determinantes == '?':
                        determinantes = 'm,'
                    else: 
                        determinantes = determinantes + 'm,'
                    dispatcher.utter_message(text="Hay forma inanimada(determinante)")
                else:
                    dispatcher.utter_message(text="No hay forma(determinante)")
        if slot_movimiento == "true" and slot_forma == 'humana':
            if determinantes == '?':
                determinantes = 'M,'
            else: 
                determinantes = determinantes + 'M,'
            dispatcher.utter_message(text="Hay movimiento humano M(determinante)")
        else:
            if slot_movimiento == "true" and slot_forma == 'inanimada':
                if determinantes == '?':
                    determinantes = 'm,'
                else: 
                    determinantes = determinantes + 'm,'
                dispatcher.utter_message(text="Hay movimiento inanimado m(determinante)")
            elif slot_movimiento == "true":
                if determinantes == '?':
                    determinantes = 'ind,'
                else: 
                    determinantes = determinantes + 'ind,'
                dispatcher.utter_message(text="Hay movimiento indefinido(determinante)")
            else:
                dispatcher.utter_message(text="No hay movimiento(determinante)")
        if slot_textura == "true":
            if determinantes == '?':
                determinantes = 'T,'
            else: 
                determinantes = determinantes + 'T,'
            dispatcher.utter_message(text="Hay textura(determinante)")
        else:
            dispatcher.utter_message(text="No hay textura(determinante)") 
        if slot_reflejo == "true":
            if determinantes == '?':
                determinantes = 'r,'
            else: 
                determinantes = determinantes + 'r,'
            dispatcher.utter_message(text="Hay reflejo(determinante)")
        else:
            dispatcher.utter_message(text="No hay reflejo(determinante)")
        
        # Lo que hace es mostrar el mensaje con la próxima imagen: utter_Lamina2 ó utter_Lamina3
        next_response = tracker.get_slot("response")
        if next_response != "None":
            dispatcher.utter_message(response=next_response)  

        wb = op.load_workbook('PlanillaZulliger.xlsx')
        ws = wb.get_sheet_by_name('Hoja de datos')
        if determinantes != '?':
            determinantes = determinantes[:-1]
        if contenidos != '?':
            contenidos = contenidos[:-1]
        if lamina == (1 or 2 or 3):
            ws.append([lamina,'1','?','?','?', determinantes,'?', par, contenidos, popular,'?','?']) 
        else:
            if lamina == 4:
                #lamina == 1
                vf2 = ws['F2'].value
                ws['F2'] = str(vf2)  + determinantes #recortar repetidos y poner comas
                ws['H2'] = par
                vi2 = ws['I2'].value
                ws['I2'] = str(vi2) + contenidos
                vj2 = ws['J2'].value
                ws['J2'] = str(vj2) + popular 
            elif lamina == 5:
                lamina == 2
            elif lamina == 6:
                lamina == 3
            
        #ws.append([lamina,'1','?','?','?', determinantes,'?', par, contenidos, popular,'?','?']) ##necesito mergearlas no se como, o update de fila. 
        #ws.merge([lamina,'1','?','?','?', determinantes,'?', par, contenidos, popular,'?','?'])
        wb.save('PlanillaZulliger.xlsx')
        wb.close()
        planilla = pd.read_excel('PlanillaZulliger.xlsx')
        print(planilla)

        return [SlotSet("par","false"),SlotSet("vista","false"),SlotSet("color_cromatico","false"),SlotSet("color_acromatico","false"),SlotSet("forma","false"),SlotSet("movimiento","false"),SlotSet("textura","false"),SlotSet("reflejo","false"),SlotSet("response", "None")]

